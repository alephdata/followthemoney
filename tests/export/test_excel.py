import os
from unittest import TestCase
from tempfile import mkstemp
from openpyxl import load_workbook

from followthemoney import model
from followthemoney.export.excel import ExcelExporter


ENTITY = {
    'id': 'person',
    'schema': 'Person',
    'properties': {
        'name': 'Ralph Tester',
        'birthDate': '1972-05-01',
        'idNumber': ['9177171', '8e839023'],
        'website': 'https://ralphtester.me',
        'phone': '+12025557612',
        'email': 'info@ralphtester.me'
    }
    }


class ExcelExportTestCase(TestCase):

    def setUp(self):
        _, self.temp = mkstemp(suffix='.xlsx')

    def tearDown(self):
        os.unlink(self.temp)

    def test_excel_export(self):
        entity = model.get_proxy(ENTITY)
        exporter = ExcelExporter(self.temp, extra=['source'])
        exporter.write(entity, extra=['test'])
        exporter.finalize()
        workbook = load_workbook(self.temp)
        self.assertListEqual(workbook.sheetnames, ['People'])
        sheet = workbook["People"]
        rows = list(sheet)
        props = exporter.exportable_properties(entity.schema)
        self.assertListEqual(
            [cell.value for cell in rows[0]],
            ['ID', 'source'] +
            [prop.label for prop in props]
        )
        self.assertListEqual(
            [cell.value for cell in rows[1][:3]],
            ['person', 'test', 'Ralph Tester']
        )

    def test_excel_bytesio(self):
        entity = model.get_proxy(ENTITY)
        exporter = ExcelExporter(self.temp, extra=['source'])
        exporter.write(entity, extra=['test'])
        buffer = exporter.get_bytesio()
        assert len(buffer.getvalue()) > 100
