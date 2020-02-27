import logging
from io import BytesIO
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.exceptions import IllegalCharacterError

from followthemoney.export.common import Exporter
from followthemoney.util import sanitize_text

log = logging.getLogger(__name__)


class ExcelWriter(object):
    HEADER_FONT = Font(bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill(start_color='982022',
                              end_color='982022',
                              fill_type='solid')

    def __init__(self):
        self.workbook = Workbook(write_only=True)

    def make_sheet(self, title, headers):
        sheet = self.workbook.create_sheet(title=title)
        sheet.freeze_panes = 'A2'
        sheet.sheet_properties.filterMode = True
        cells = []
        for header in headers:
            header = sanitize_text(header)
            cell = WriteOnlyCell(sheet, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cells.append(cell)
        sheet.append(cells)
        return sheet

    def get_bytesio(self):
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer


class ExcelExporter(ExcelWriter, Exporter):

    def __init__(self, file_path, extra=None):
        ExcelWriter.__init__(self)
        Exporter.__init__(self)
        self.file_path = file_path
        self.extra = extra or []
        self.sheets = {}

    def write(self, proxy, extra=None, **kwargs):
        if proxy.schema not in self.sheets:
            headers = ['ID']
            headers.extend(self.extra)
            for prop in self.exportable_properties(proxy.schema):
                headers.append(prop.label)
            sheet = self.make_sheet(proxy.schema.plural, headers)
            self.sheets[proxy.schema] = sheet
        sheet = self.sheets.get(proxy.schema)
        try:
            cells = [proxy.id]
            cells.extend(extra or [])
            for prop, values in self.exportable_fields(proxy):
                cells.append(prop.type.join(values))
            sheet.append(cells)
        except IllegalCharacterError as ice:
            log.error("Invalid text for Excel export: %s", ice)

    def finalize(self):
        self.workbook.save(self.file_path)
